"""Interface to read tabular data stored in an Excel 2007- (xlsx) workbook file.

*** Version History:

2019-07-23, first functional version
2019-07-27, change to subclass of Table
"""
# **** general test for unique subjectID here, or in calling module ??? *********

# handle record as dict for simpler recode_fcn and accept_row  ??? *******

# ***** define columns by column address or also by column index or column header ??? **********

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import get_column_letter


from ItemResponseCalc import table_reader

# import logging
# logger = logging.getLogger(__name__)


class FileFormatError(RuntimeError):
    """Format error causing non-usable data"""

#
# class ParameterError(pc_base.FileReadError):
#     """Error in calling parameters causing non-usable data"""


class ArgumentError(RuntimeError):
    """Error in calling parameters causing non-usable data"""


# ------------------------------------------------------------------------

class Table(table_reader.Table):
    """Interface to ONE sheet of an Excel xlsx workbook storing tabular data
    """
    def __init__(self, file_path,
                 fields=None,
                 sheet=None,
                 header_row=1,
                 **kwargs):
        """
        :param file_path: Path to file for reading
        :param fields: list of fields to be returned in each record
            if None, ALL fields are returned
        :param sheet: (optional) sheet name.
            If undefined, only the first work-book sheet is used.
        :param header_row: row with header labels for all columns
            if header_row is None or == 0, column addresses are used as field names
        :param kwargs: additional properties for super-class,
            field_types, keyed, accept_row, missing_values, recode_fcns
        """
        self.sheet = sheet
        if header_row is None:
            header_row = 0
        self.header_row = header_row
        super().__init__(table_id=file_path, fields=fields, **kwargs)

    def preview(self, max_records=10, keyed=False):
        """Read headers and some raw data rows from the file,
        as stored without any type casting or other checks
        :param max_records: limit on the number of records
        :param keyed: boolean: if True, return data as dict # *********************** ???
        :return: iterator over data rows, as stored without any type casting or other checks
        """
        wb = self._get_workbook()
        ws = self._get_sheet(wb)
        # result = []
        rows = ws.iter_rows(min_row=1)
        for (row_n, row) in enumerate(rows):
            if row_n > max_records:
                break
            # result.append([cell.value for cell in row])
            yield [cell.value for cell in row]
        wb.close()
        # return result

    def __iter__(self):
        """Open file for read-only access,
        yielding one row at a time,
        including only selected fields,
        selected, recoded, and pre-processed by self.process_row().
        :return: iterator object over file rows,
            starting just after self.header_row
        """
        wb = self._get_workbook()
        ws = self._get_sheet(wb)
        header = self._get_header(ws)
        if self.fields is None:
            fields_index = None
        else:
            fields_index = [header.index(f) for f in self.fields]
            # ************** must catch ValueError if f is not in header
        rows = ws.iter_rows(min_row=self.header_row + 1)
        for (row_n, row) in enumerate(rows):
            if row_n % self.sample_factor == 0:
                if fields_index is None:
                    row_val = [r.value for r in row]
                else:
                    row_val =[row[i].value for i in fields_index]
                row_val = self.process_row(row_val)
                if row_val is not None:
                    yield row_val
        wb.close()  # ******** might never get here ******** need context manager ?

    def __len__(self):
        """Number of rows in file, except header
        :return: scalar integer
        """
        wb = self._get_workbook()
        ws = self._get_sheet(wb)
        n = ws.max_row - self.header_row
        wb.close()
        return n // self.sample_factor

    # ------------------------------------------------ Private:
    def _get_workbook(self):
        """Open work-book and get ONE selected sheet
        :return: ws = a work-sheet object
        """
        try:
            return load_workbook(str(self.table_id), read_only=True)
        except InvalidFileException:
            raise FileFormatError(f'Cannot load workbook from file {self.table_id.stem}')

    def _get_sheet(self, wb):
        """Get ONE selected sheet from an open workbook
        :param wb: active WorkBook object
        :return: ws = a work-sheet object
        """
        if self.sheet is None:
            self.sheet = wb.sheetnames[0]
        try:
            return wb[self.sheet]
        except IndexError:
            raise FileFormatError(f'Cannot load sheet {repr(self.sheet)} from file {self.table_id.stem}')

    def _get_header(self, ws):
        """get file header from the worksheet, if it has a header row
        :param ws: selected WorkSheet of a Workbook open for read-only access
        :return: list with header name strings from the file
            or column addresses if there is no header row
        """
        if self.header_row == 0:
            row1 = ws[1]
            return [get_column_letter(i) for i in range(1, 1+len(row1))]
        else:
            header_row = ws[self.header_row]
            return [cell.value for cell in header_row]

# --------------------------------------------------- help sub-functions


# -----------------------------------------------------------------------------
# if __name__ == '__main__':
#     from pathlib import Path
#
#     HAQ_PATH = Path.home() / 'Documents/LeijonKTH/heartech/HA-QualityRegistry'
#     # top-dir for everything to be used here
#
#     ioiha_data_path = HAQ_PATH / 'IRT-IOI-HA' / 'IOI-HA-data'
#     # to ioi-ha data sets OTHER THAN the Swedish Quality Registry
#
#     ioiha_hickson = ioiha_data_path / 'Hickson'
#
#     hickson_path = ioiha_hickson / 'Short_Eartrak AUS.xlsx'
#
#     table = Table(file_path=hickson_path,
#                   fields='CDEFGHI',
#                   header_row=0,
#                   missing_values=['.'])
#
#     print('\n*** Using table(header_row=0): \n')
#     print('Number of rows = ', len(table))
#     for (row_n, row) in enumerate(table):
#         if row_n > 11:
#             break
#         print(f'row #{row_n} = {row}')
#
#     print('table.preview(10)=\n', table.preview(10))
#     for row in table.preview(10):
#         print(row)
#
#     table = Table(file_path=hickson_path,
#                   fields=['Q01', 'Q02', 'Q03', 'Q04', 'Q05', 'Q06', 'Q07'],
#                   header_row=1,
#                   missing_values=['.'])
#
#     print('\n*** Using table(header_row=1): \n')
#     print('Number of rows = ', len(table))
#     for (row_n, row) in enumerate(table):
#         if row_n > 10:
#             break
#         print(f'row #{row_n} = {row}')
#
    # print('Hickson file.preview(10)=\n',table.preview(10))

