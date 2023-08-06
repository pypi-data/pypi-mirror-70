# encoding=utf8
from openpyxl.styles import Alignment, Side, Border
from openpyxl import Workbook
import string
import os
import time


def create_sheet(info, model, headers, props_names, data):
    path = "/tmp/export"
    time_str = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    name = f"{info.model['model']}_{time_str}.xlsx"
    file_name = f"{path}/{name}"

    if not os.path.exists(path):  # 如果路径不存在
        os.makedirs(path)

    wb = Workbook(write_only=True)
    work_sheet = wb.create_sheet(model, 0)
    # 水平居中, 垂直居中
    alignment_style = Alignment(vertical='center')
    # work_sheet['A1'].font = Font(size=16, bold=True)
    # work_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 定义Border边框样式
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    border_style = Border(left=left, right=right, top=top, bottom=bottom)
    # work_sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(head_list))
    # work_sheet.cell(row=1, column=1).value = title

    # 头部标题
    # for index, item in enumerate(headers):
    #     col = index + 1
    #     work_sheet.cell(row=1, column=col).value = item
    #     # 设置单元格样式文本水平居中、垂直居中、设置字体加粗、背景颜色、边框样式
    #     work_sheet.cell(row=1, column=col).alignment = alignment_style
    #     work_sheet.cell(row=1, column=col).font = Font(bold=True, size=9)
    #     work_sheet.cell(row=1, column=col).fill = PatternFill(fill_type='solid', fgColor='EE9A49')
    #     work_sheet.cell(row=1, column=col).border = border_style
    #     # 设置列宽、生成前26大写字母  ascii_uppercase生成所有大写字母

    # 头部标题
    work_sheet.append(headers)

    upper_string = string.ascii_uppercase[:26]
    for col in upper_string:
        work_sheet.column_dimensions[col].width = 20

    # 数据
    # row = 2
    # for line in data:
    #     for index, item in enumerate(props_names):
    #         col = index + 1
    #         work_sheet.cell(row, column=col).value = line[item]
    #         # 设置单元格样式文本水平居中、垂直居中、设置边框样式、设置字体大小
    #         work_sheet.cell(row, column=col).alignment = alignment_style
    #         work_sheet.cell(row, column=col).border = border_style
    #         work_sheet.cell(row, column=col).font = Font(size=9)
    #     row = row + 1

    # 数据
    for line in data:
        row = []
        for index, item in enumerate(props_names):
            row.append(line[item])
        work_sheet.append(row)

    wb.save(file_name)
    wb.close()

    return path, name
