#!/usr/bin/env python
"""
CGNX API -> Machines Serial Report

cloudgenix_serial_report@ebob9.com

"""
# standard modules
import datetime
import os
import argparse

# extra modules
import cloudgenix
import cloudgenix_idname
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter


DATE_FMT = "%D %T"

HEADER = [
    'Serial Number',
    'Model',
    'Currently Connected',
    'Creation Date (UTC)',
    'Hardware State',
    'Claimed to Element',
    'Claimed Date (UTC)',
    'Element State',
    'Element Name',
    'Element Software Version',
    'Site'
]

SCRIPT_VERSION = cloudgenix.version
SCRIPT_NAME = 'CloudGenix Serial Report'

# set date style
date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY HH:MM:MM')


def generate(logged_in_sdk, specify_filename=None):

    # get time now.
    curtime_str = datetime.datetime.utcnow().strftime('%Y-%m-%d-%H-%M-%S')

    # create file-system friendly tenant str.
    tenant_str = "".join(x for x in logged_in_sdk.tenant_name if x.isalnum()).lower()

    # Set filename
    if specify_filename is None:
        report_xlsx = os.path.join('./', '{0}_serials_{1}.xlsx'.format(tenant_str, curtime_str))
    else:
        report_xlsx = specify_filename

    # Get names map
    print("Building ID->Name lookup table.")
    id_name, _, _, _ = cloudgenix_idname.siteid_to_name_dict(logged_in_sdk)

    # get machines and elements
    print("Querying APIs.")
    machines_response = logged_in_sdk.get.machines()
    elements_response = logged_in_sdk.get.elements()

    report = []

    print("Parsing Responses.")
    if machines_response.cgx_status and elements_response.cgx_status:
        machines_list = machines_response.cgx_content.get('items', [])
        elements_list = elements_response.cgx_content.get('items', [])

        element_lookup = {}
        # create elements lookup table
        for element in elements_list:
            entry = {}
            el_serial_number = element.get('serial_number')
            el_state = element.get('state', '<Unknown>')
            el_name = element.get('name', '<Unnamed>')
            el_software_version = element.get('software_version', '<Unknown>')
            el_claimed_date = element.get('_created_on_utc')
            el_site_id = element.get('site_id')

            if el_serial_number:
                entry['state'] = el_state
                entry['name'] = el_name
                entry['version'] = el_software_version
                if el_claimed_date and isinstance(el_claimed_date, int) and el_claimed_date > 0:
                    claimed_datetime = datetime.datetime.utcfromtimestamp(int(el_claimed_date) / 10000000.0)
                    entry['claimed'] = claimed_datetime
                else:
                    entry['claimed'] = "<Unknown>"

                if el_site_id:
                    entry['site'] = id_name.get(el_site_id, "<Site Name Unavaialble: {0}>".format(el_site_id))
                else:
                    entry['site'] = "<Not used or Unknown>"

                # add to lookup
                element_lookup[el_serial_number] = entry

        # Build report.
        for machine in machines_list:
            report_entry = []
            m_serial_number = machine.get('sl_no')
            m_model_name = machine.get('model_name', "<Unknown>")
            m_connected = machine.get('connected', "<Unknown>")
            m_state = machine.get('machine_state', "<Unknown>")
            m_created = machine.get('_created_on_utc')

            if m_serial_number:
                # serial
                report_entry.append(m_serial_number)
                # model
                report_entry.append(m_model_name)
                # connected
                if isinstance(m_connected, bool):
                    report_entry.append('Yes' if m_connected else 'No')
                else:
                    report_entry.append(m_connected)
                # creation date
                if m_created and isinstance(m_created, int) and m_created > 0:
                    created_datetime = datetime.datetime.utcfromtimestamp(int(m_created) / 10000000.0)
                    report_entry.append(created_datetime)
                else:
                    report_entry.append("<Unknown>")

                # HW state
                report_entry.append(m_state)

                # Check if this machine is an element
                elem_data = element_lookup.get(m_serial_number)
                if elem_data:
                    claimed = 'Yes'
                    claimed_date = elem_data.get('claimed')
                    state = elem_data.get('state')
                    name = elem_data.get('name')
                    version = elem_data.get('version')
                    site = elem_data.get('site')
                else:
                    claimed = 'No'
                    claimed_date = ""
                    state = ""
                    name = ""
                    version = ""
                    site = ""

                # add element rows
                # claimed to element
                report_entry.append(claimed)
                # claimed date
                report_entry.append(claimed_date)
                # ele state
                report_entry.append(state)
                # ele name
                report_entry.append(name)
                # version
                report_entry.append(version)
                # site
                report_entry.append(site)

                # add report entry to report
                report.append(report_entry)

    print("Building Reports.")
    # build report
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # create header
    ws.append(HEADER)

    # bold header
    for cell in ws['1:1']:
        cell.font = Font(bold=True)

    # Build report
    for row in report:
        ws.append(row)

    # autofit columns to data, set date style on date columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        column_value = col[0].value

        for cell in col:
            if isinstance(cell.value, datetime.datetime):
                cell_check_len = len(str(cell.value.strftime(DATE_FMT)))
                # date time cells, set format
                cell.style = date_style
            else:
                cell_check_len = len(str(cell.value))
            if cell_check_len > max_length:
                max_length = cell_check_len

        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width

    # enable filters
    ws.auto_filter.ref = "A1:{0}{1}".format(get_column_letter(ws.max_column), ws.max_row)

    # title from file name
    title = os.path.basename(report_xlsx).rstrip('.xlsx')
    # remove invalid chars
    for char in "\\/*[]:?":
        title = title.replace(char, '')
    # only use first 30 chars of title
    ws.title = title[:30]

    # Save the file
    wb.save(report_xlsx)


def go():

    ############################################################################
    # Begin Script, start login / argument handling.
    ############################################################################

    # Parse arguments
    parser = argparse.ArgumentParser(description="{0}.".format(SCRIPT_NAME))

    # Allow Controller modification and debug level sets.
    controller_group = parser.add_argument_group('API', 'These options change how this program connects to the API.')
    controller_group.add_argument("--controller", "-C",
                                  help="Controller URI, ex. https://api.elcapitan.cloudgenix.com",
                                  default=None)

    controller_group.add_argument("--insecure", "-I", help="Disable SSL certificate and hostname verification",
                                  dest='verify', action='store_false', default=True)

    controller_group.add_argument("--noregion", "-NR", help="Ignore Region-based redirection.",
                                  dest='ignore_region', action='store_true', default=False)

    login_group = parser.add_argument_group('Login', 'These options allow skipping of interactive login')
    login_group.add_argument("--email", "-E", help="Use this email as User Name instead of prompting",
                             default=None)
    login_group.add_argument("--pass", "-PW", help="Use this Password instead of prompting",
                             default=None)

    debug_group = parser.add_argument_group('Debug', 'These options enable debugging output')
    debug_group.add_argument("--debug", "-D", help="Verbose Debug info, levels 0-2", type=int,
                             default=0)

    output_group = parser.add_argument_group('Output', 'These options change how the output is generated.')
    output_group.add_argument("--output", help="Output file name (default is auto-generated from name/date/time)",
                              type=str, default=None)

    args = vars(parser.parse_args())

    ############################################################################
    # Instantiate API
    ############################################################################

    sdk = cloudgenix.API(controller=args["controller"], ssl_verify=args["verify"])

    # set debug
    sdk.set_debug(args["debug"])

    # set ignore region
    sdk.ignore_region = args["ignore_region"]

    ############################################################################
    # Draw Interactive login banner, run interactive login including args above.
    ############################################################################

    print("{0} v{1} ({2})\n".format(SCRIPT_NAME, SCRIPT_VERSION, sdk.controller))

    # interactive or cmd-line specified initial login

    while sdk.tenant_name is None:
        sdk.interactive.login(args["email"], args["pass"])

    ############################################################################
    # End Login handling, begin script..
    ############################################################################

    generate(sdk, specify_filename=args["output"])


if __name__ == "__main__":
    go()
