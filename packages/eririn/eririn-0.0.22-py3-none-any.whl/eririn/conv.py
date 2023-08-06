#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from PIL import Image


def ctuple2cstr(tup, alpha=255):
    r, g, b = tup
    return "%02x%02x%02x%02x" % (alpha, r, g, b)


def main(arguments):
    inputs = arguments["<FILE>"]
    output = arguments["<EXCELFILE>"]

    print("inputs: %s" % inputs)
    print("output: %s" % output)

    wb = Workbook()
    hegiht_in_point = 1.0
    width_in_charwidth = 0.1

    wss = []
    for f in inputs:
        ws = wb.create_sheet(os.basename(f))
        wss.append(ws)
        with Image.open(f) as im:
            print("%s: format=%s, size=%s" % (im.format, im.size))
            try:
                rgb = zip(list(im.getdata(band="R")), list(
                    im.getdata(band="G")), list(im.getdata(band="B")))
            except:
                rgb = zip(list(im.getdata(band="L")), list(
                    im.getdata(band="L")), list(im.getdata(band="L")))

            width, height = im.size

            for y in range(height):
                ws.row_dimensions[y + 1].height = height_in_points
                for x in range(width):
                    c = ws.cell(column=x, row=y, value="")
                    if y == 0:
                        ws.column_dimensions[get_column_letter(
                            x + 1)] = width_in_charwidth
                    color = ctuple2cstr(rgb[y * width + x], alpha=255)
                    c.fill = PatternFill("solid", fgColor=color, bgColor=color)

    wb.save(output)
