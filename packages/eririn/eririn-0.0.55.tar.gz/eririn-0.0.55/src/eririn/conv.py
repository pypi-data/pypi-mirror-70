#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter
from PIL import Image
from tqdm import tqdm


def ctuple2cstr(tup, alpha=0xff):
    r, g, b = tup
    return ("%02x%02x%02x" % (r, g, b)).upper()


def main(arguments):
    inputs = arguments["<FILE>"]
    output = arguments["<EXCELFILE>"]

    print("inputs: %s" % inputs)
    print("output: %s" % output)

    wb = Workbook()
    height_in_points = 10.0
    width_in_charwidth = 1.0

    wss = []
    for fidx, f in enumerate(inputs):
        if fidx == 0:
            ws = wb.active
            ws.title = os.path.basename(f)
        else:
            ws = wb.create_sheet(os.path.basename(f))

        with Image.open(f) as im:
            print("%s: format=%s, size=%s" % (f, im.format, im.size))
            bands = im.getbands()
            cdict = {}
            for ix, b in enumerate(bands):
                if b == "R":
                    cdict["R"] = ix
                elif b == "G":
                    cdict["G"] = ix
                elif b == "B":
                    cdict["B"] = ix
                elif b == "L":
                    cdict["L"] = ix
                elif b == "A":
                    cdict["A"] = ix
            try:
                rgb = zip(list(im.getdata(band=cdict["R"])), list(
                    im.getdata(band=cdict["G"])), list(im.getdata(band=cdict["B"])))
            except:
                rgb = zip(list(im.getdata(band="L")), list(
                    im.getdata(band="L")), list(im.getdata(band="L")))

            rgb = list(rgb)

            width, height = im.size
            c_width, c_height = int(width * 0.6), int(height * 0.6)

            for x in tqdm(range(c_width)):
                for y in range(c_height):
                    _ = ws.cell(column=(x + 1), row=(y + 1), value=" ")
                    color = ctuple2cstr(rgb[y * width + x])
                    c = ws[get_column_letter(x + 1) + ("%d" % (y + 1))]
                    # cl = Color(rgb=color, type="rgb")
                    # c.font = Font(name='Calibri', size=11, bold=False, italic=False,
                    #     vertAlign=None, underline='none', strike=False, color=color)
                    c.fill = PatternFill(fgColor=color, fill_type="solid")

            for y in range(c_height):
                ws.row_dimensions[y + 1].height = height_in_points

            for x in range(c_width):
                ws.column_dimensions[get_column_letter(
                    x + 1)].width = width_in_charwidth

    wb.save(output)
