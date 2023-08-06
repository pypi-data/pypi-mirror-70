#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from operator import itemgetter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter
from PIL import Image
from tqdm import tqdm

# LIMIT_OF_EXCEL = 0x18001
LIMIT_OF_EXCEL = 0x18001 * 2


def ctuple2cstr(tup, dic=None):
    if dic is not None:
        tup = dic[tup]

    r, g, b = tup

    # silly clamp
    if r > 255:
        r = 255
    if g > 255:
        g = 255
    if b > 255:
        b = 255

    return "%02X%02X%02X" % (r, g, b)


def resize_tuple(width, height):
    all_pixel = width * height
    if all_pixel < LIMIT_OF_EXCEL:
        return (width, height)
    alpha = (LIMIT_OF_EXCEL / all_pixel) ** 0.5
    print("alpha = %f" % alpha)
    r_width, r_height = int(width * alpha), int(height * alpha)
    actual_pixel = r_width * r_height
    print("actual pixel to write = %d" % actual_pixel)

    return (r_width, r_height)


def main(arguments):
    inputs = arguments["<FILE>"]
    output = arguments["<EXCELFILE>"]

    print("inputs: %s" % inputs)
    print("output: %s" % output)

    wb = Workbook()
    height_in_points = 10.0
    width_in_charwidth = 1.6

    wss = []
    for fidx, f in enumerate(inputs):
        if fidx == 0:
            ws = wb.active
            ws.title = os.path.basename(f)
        else:
            ws = wb.create_sheet(os.path.basename(f))

        with Image.open(f) as im:
            width, height = im.size
            resized = resize_tuple(width, height)
            c_width, c_height = resized
            im2 = im.resize(resized, resample=Image.LANCZOS)
            print("%s: format=%s, size=%s" % (f, im.format, im.size))
            if c_width != width or c_height != height:
                print("resized to: size=%s" % str(im2.size))
            bands = im2.getbands()
            cdict = {}
            for ix, b in enumerate(bands):
                if b == "R":
                    cdict["R"] = ix
                elif b == "G":
                    cdict["G"] = ix
                elif b == "B":
                    cdict["B"] = ix
                elif b == "L":
                    cdict["L"] = ix
                elif b == "A":
                    cdict["A"] = ix
            try:
                rgb = zip(list(im2.getdata(band=cdict["R"])), list(
                    im2.getdata(band=cdict["G"])), list(im2.getdata(band=cdict["B"])))
            except:
                rgb = zip(list(im2.getdata(band="L")), list(
                    im2.getdata(band="L")), list(im2.getdata(band="L")))

            rgb = list(rgb)
            rgbmap = {(r, g, b): (r, g, b) for r, g, b in rgb}
            num_c = len(list(rgbmap.keys()))
            reduced_ratio = 0.5
            spl = int((num_c * reduced_ratio) ** 0.3333333333333333)
            spl1p = spl + 1
            st = int(256 / spl)

            index_dict = {}
            for i in range(spl1p):
                for j in range(spl1p):
                    for k in range(spl1p):
                        index_dict[i * spl1p * spl1p + j * spl1p + k] = 0

            new_color_dict = {}
            for idx, s in enumerate(list(rgbmap.keys())):
                r, g, b = s
                mr, mg, mb = int(
                    ((r + 0.5) / 256) * 256), int(((g + 0.5) / 256) * 256), int(((b + 0.5) / 256) * 256)
                i, j, k = int(r / 256 * st), int(g /
                                                 256 * st), int(b / 256 * st)
                index = i * spl1p * spl1p + j * spl1p + k
                if index_dict[index] == 0:
                    index_dict[index] = (
                        int((r / st + 0.5) * st), int((g / st + 0.5) * st), int((b / st + 0.5) * st))
                new_color_dict[s] = index_dict[index]

            bytes_written = 0
            for x in tqdm(range(c_width)):
                for y in range(c_height):
                    _ = ws.cell(column=(x + 1), row=(y + 1), value=" ")
                    bytes_written += 1
                    if bytes_written > LIMIT_OF_EXCEL:
                        break
                    color = ctuple2cstr(
                        rgb[y * c_width + x], dic=new_color_dict)
                    c = ws[get_column_letter(x + 1) + ("%d" % (y + 1))]
                    c.fill = PatternFill(fgColor=color, fill_type="solid")

            for y in range(c_height):
                ws.row_dimensions[y + 1].height = height_in_points

            for x in range(c_width):
                ws.column_dimensions[get_column_letter(
                    x + 1)].width = width_in_charwidth

    wb.save(output)
