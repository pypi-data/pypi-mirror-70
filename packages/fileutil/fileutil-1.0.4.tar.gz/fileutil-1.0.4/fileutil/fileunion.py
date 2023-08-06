from pathlib import Path
import openpyxl
import csv
from . import EXCEL_TYPE, TXT_TYPE
from . import common
from .filetransform import excel_to_csv
from .parser import union_file


def union_in_path_to_csv(save_path, file_path, file_type=EXCEL_TYPE, all=False, engine='normal', skip_row=0,
                         need_title=False, **kwargs):
    save_path, file_path = Path(save_path), Path(file_path)
    if not save_path.exists():
        raise ValueError('保存路径不存在')
    if not file_path.exists():
        raise ValueError('文件搜索路径不存在')
    file_list = common.get_files(file_path, file_type, all)
    need_trans_list_before = list(filter(lambda f: f.suffix in EXCEL_TYPE, file_list))
    no_need_trans_list = [x for x in file_list if x not in need_trans_list_before]
    need_trans_list_after = []
    kwargs.update(skip_row=skip_row, need_title=need_title, file_type=file_type, all=all)
    for file in need_trans_list_before:
        file = Path(file)
        need_trans_list_after += excel_to_csv(file.parent, file, **kwargs)
    need_union_file = need_trans_list_after + no_need_trans_list
    result = union_file(save_path, need_union_file, engine=engine, **kwargs)
    [f.unlink() for f in need_trans_list_after]  # 删除转换时的临时文件
    return result


def union_in_list_to_csv(save_path, file_list, engine='normal', skip_row=0, need_title=False, **kwargs):
    save_path = Path(save_path)
    if not save_path.exists():
        raise ValueError('保存路径不存在')
    need_trans_list_before = list(filter(lambda f: f.suffix in EXCEL_TYPE, file_list))
    no_need_trans_list = [x for x in file_list if x not in need_trans_list_before]
    need_trans_list_after = []
    kwargs.update(skip_row=skip_row, need_title=need_title)
    for file in need_trans_list_before:
        file = Path(file)
        need_trans_list_after += excel_to_csv(file.parent, file, **kwargs)
    need_union_file = need_trans_list_after + no_need_trans_list
    result = union_file(save_path, need_union_file, engine=engine, **kwargs)
    [f.unlink() for f in need_trans_list_after]  # 删除转换时的临时文件
    return result


def union_sheets_to_csv(save_path, file_path, engine='normal', skip_row=0, need_title=False, **kwargs):
    file_path = Path(file_path)
    kwargs.update(skip_row=skip_row, need_title=need_title)
    csv_list = excel_to_csv(save_path, file_path, **kwargs)
    result = union_in_list_to_csv(file_path.parent, csv_list, engine=engine, **kwargs)
    [f.unlink() for f in csv_list]
    return result


def union_csvs_in_list_to_sheets(save_path, file_list, **kwargs):
    if not save_path.parent.exists():
        raise ValueError('保存路径不存在')
    if save_path.exists():
        dwb = openpyxl.load_workbook(save_path)
        need_delete_sheet = None
    else:
        dwb = openpyxl.Workbook()
        need_delete_sheet = dwb.active
    result = []
    for file in file_list:
        encode = common.file_encoding(file)
        fobj = open(file, 'r', encoding=encode)
        dws = dwb.create_sheet(title=file.stem)
        lines = csv.reader(fobj)
        for row_index, row in enumerate(lines):
            for column_index, cell in enumerate(row):
                dws.cell(row=row_index + 1, column=column_index + 1, value=cell)
        print(f'合并文件：{file}')
        result.append(file)
    if need_delete_sheet:
        dwb.remove(need_delete_sheet)
    dwb.save(save_path)
    return result


def union_excels_in_list_to_sheets(save_path, file_list, **kwargs):
    if not save_path.parent.exists():
        raise ValueError('保存路径不存在')
    if save_path.exists():
        dwb = openpyxl.load_workbook(save_path)
        need_delete_sheet = None
    else:
        dwb = openpyxl.Workbook()
        need_delete_sheet = dwb.active
    result = []
    for file in file_list:
        swb = openpyxl.load_workbook(file)
        for sws in swb.worksheets:
            dws = dwb.create_sheet(title=sws.title)
            for i, row in enumerate(sws.iter_rows()):
                for j, cell in enumerate(row):
                    dws.cell(row=i + 1, column=j + 1, value=cell.value)
            print(f'合并文件：{file}[{sws.title}]')
        result.append(file)
    if need_delete_sheet:
        dwb.remove(need_delete_sheet)
    dwb.save(save_path)
    return result


def union_in_path_to_sheets(save_path, file_path, file_type=EXCEL_TYPE, all=False, **kwargs):
    ave_path, file_path = Path(save_path), Path(file_path)
    if not save_path.parent.exists():
        raise ValueError('保存路径不存在')
    if not file_path.exists():
        raise ValueError('文件搜索路径不存在')
    if save_path.exists():
        save_path.unlink()
    file_list = common.get_files(file_path, file_type, all)
    result = union_in_list_to_sheets(save_path, file_list)
    return result


def union_in_list_to_sheets(save_path, file_list, **kwargs):
    if not save_path.parent.exists():
        raise ValueError('保存路径不存在')
    result = []
    excel_files = [f for f in file_list if f.suffix in EXCEL_TYPE]
    if len(excel_files) > 0:
        result = union_excels_in_list_to_sheets(save_path, excel_files)
    txt_files = [f for f in file_list if f.suffix in TXT_TYPE]
    if len(txt_files) > 0:
        result += union_csvs_in_list_to_sheets(save_path, txt_files)
    return result
